import openpyxl
import os, sys
from openpyxl import Workbook, load_workbook

lt_excelData = [["ID", "Name", "Age", "Country"],
                [1   , "Suares", 33 , "URQ"    ],
                [2   , "Kean"  , 22 , "ENG"    ],
                [3   , "Mbappe", 19 , "FRA"    ],
                [4,  "Neymar",  28,       "BRA"],
                [5,  "Messi",  34,        "ARG"],
                [6,  "Quang Hai",  23,    "VN" ],
                [7,  "Ronaldo",  35,      "POR"],
                [8,  "Xuan Truong",  23,  "VN" ],
                [9,  "Cong Phuong",  23,  "VN" ],
                [10, "Tuan Anh",  23,     "CAM"]]
lt_insData = [[11, "Hazard",   33, "BEL"],
              [22, "Coutinho", 22, "BRA"]]

if(len(sys.argv) < 2):
    print ("missing input excel file name !")
else:
    print ("process !")
    excelfile = sys.argv[1]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Excel_DB"
    #ws1 = wb.create_sheet("excel_db",0)
    num_row = len(lt_excelData[:])
    num_col = len(lt_excelData[0][:])
    print ("number of row :",num_row)
    print ("number of column :",num_col)

    # 1. Insert all above data to an excel file
    for row in range(len(lt_excelData[:])):
        ws1.append(lt_excelData[row])

    # 2. Update Country of Tuan Anh to VN
    #ws1["D11"] = "VN"
    #ws1.cell(row=11, column=4, value = "VN")
    for row in ws1.iter_rows(min_row=1, max_row=11, min_col=1, max_col=4):
        if(row[1].value == "Tuan Anh"):
            row[3].value = "VN"

    # 3. Insert 2 items below to the above excel file :
    for row in range(len(lt_insData[:])):
        ws1.append(lt_insData[row])

    wb.save(excelfile+".xlsx")

    wb = load_workbook(excelfile+".xlsx")
    sheets = wb.sheetnames
    print (sheets)
    wslist = wb.worksheets
    print ("list of worksheet obj", wslist)
    ws2 = wb["Excel_DB"]
    ws2.delete_rows(1,11)
    wb.save(excelfile+".xlsx")

#https://www.freecodecamp.org/news/a-practical-guide-to-learning-front-end-development-for-beginners-da6516505e41/